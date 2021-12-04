from openpyxl import load_workbook
import sys
from product import ProductList
from copy import copy

COL_ID_NAME = 1
COL_ID_CODE = 2
COL_ID_TITLE = 6
COL_ID_PRICE = 10

input_file = sys.argv[1]
filter_column = 0
filter_text = ''
if len(sys.argv) >= 4:
    filter_column = int(sys.argv[2])
    filter_text = sys.argv[3]

perform_filter = filter_column>0

wb = load_workbook(input_file)
sheet = wb.active

num_rows = sheet.max_row
products = ProductList(perform_filter=perform_filter, filter_text=filter_text)

for r in range(2, num_rows+1):
    code = sheet.cell(row=r, column=COL_ID_CODE)
    name = sheet.cell(row=r, column=COL_ID_NAME)
    title = sheet.cell(row=r, column=COL_ID_TITLE)
    price = sheet.cell(row=r, column=COL_ID_PRICE)

    filter_value = ''
    if perform_filter:
        filter_value = sheet.cell(row=r, column=filter_column).value

    products.insert_product(code.value, name.value, title.value, price.value, filter_value)

sheet_name = 'Products'
if perform_filter:
    sheet_name += f' {filter_text}'
    sheet_name = sheet_name.replace('/', '-')
result_sh = wb.create_sheet(sheet_name)

RES_COL_ID_NAME = 1
RES_COL_ID_CODE = 2
RES_COL_ID_TITLE = 3
RES_COL_ID_CNT = 4
RES_COL_ID_PRICE = 5
RES_COL_ID_TOTAL_PRICE = 6

result_sh.cell(row=1, column=RES_COL_ID_NAME).value = 'عنوان گروه'
result_sh.cell(row=1, column=RES_COL_ID_CODE).value = 'کد محصول'
result_sh.cell(row=1, column=RES_COL_ID_TITLE).value = 'عنوان'
result_sh.cell(row=1, column=RES_COL_ID_CNT).value = 'تعداد'
result_sh.cell(row=1, column=RES_COL_ID_PRICE).value = 'قیمت'
result_sh.cell(row=1, column=RES_COL_ID_TOTAL_PRICE).value = 'قیمت کل'


for i in [RES_COL_ID_NAME, RES_COL_ID_CODE, RES_COL_ID_TITLE, RES_COL_ID_CNT, RES_COL_ID_PRICE, RES_COL_ID_TOTAL_PRICE]:
    result_sh.cell(row=1, column=i).font = copy(sheet.cell(row=1, column=1).font)
    result_sh.cell(row=1, column=i).fill = copy(sheet.cell(row=1, column=1).fill)


for i, p in enumerate(products.get_products()):
    row_id = i + 2
    result_sh.cell(row=row_id, column=RES_COL_ID_NAME, value=p.name)
    result_sh.cell(row=row_id, column=RES_COL_ID_CODE, value=p.code)
    result_sh.cell(row=row_id, column=RES_COL_ID_TITLE, value=p.title)
    result_sh.cell(row=row_id, column=RES_COL_ID_PRICE, value=p.price)
    result_sh.cell(row=row_id, column=RES_COL_ID_CNT, value=p.cnt)
    result_sh.cell(row=row_id, column=RES_COL_ID_TOTAL_PRICE, value=p.cnt*p.price)

result_sh.auto_filter.ref = 'A1:F1'
result_sh.sheet_view.rightToLeft = True
wb.save(input_file)
